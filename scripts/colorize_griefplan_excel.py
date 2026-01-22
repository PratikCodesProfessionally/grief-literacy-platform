import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# Farben
header_fill = PatternFill(start_color='B7E1CD', end_color='B7E1CD', fill_type='solid')  # sanftes Grün
blue_topic_fill = PatternFill(start_color='D0E7F9', end_color='D0E7F9', fill_type='solid')  # sanftes Blau
calm_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')  # sehr helles Blau
font_header = Font(bold=True, color='1B4F72')
font_topic = Font(bold=True, color='1565C0')

wb = openpyxl.load_workbook('client/public/PersonalGriefPlan_Tracking.xlsx')

for ws in wb.worksheets:
    # Header-Zeile einfärben
    ws.row_dimensions[1].height = 28
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = font_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
    # Topic-Zeile (erste Zeile nach Header) einfärben, falls vorhanden
    if ws.max_row > 1:
        for cell in ws[2]:
            cell.fill = blue_topic_fill
            cell.font = font_topic
            cell.alignment = Alignment(horizontal='center', vertical='center')
    # Restliche Zellen ruhig einfärben
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.fill = calm_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

wb.save('client/public/PersonalGriefPlan_Tracking.xlsx')
print('Excel coloring complete.')
